import pandas as pd
import openpyxl
from datetime import datetime, timedelta
import csv
import os
from collections import defaultdict

def parse_date(date_val):
    """Parse date value which could be datetime object or string in various formats"""
    if isinstance(date_val, datetime):
        return date_val
    elif isinstance(date_val, str):
        # Try different date formats
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(date_val, fmt)
            except:
                continue
    return None

def is_consecutive(date1, date2):
    """Check if two dates are consecutive (including weekends)"""
    if not date1 or not date2:
        return False
    diff = (date2 - date1).days
    return diff == 1

def group_team_periods_allowing_gaps(team_data):
    """Group periods for each team_id where the team composition does not change, allowing for gaps (e.g., weekends)."""
    if not team_data:
        return []
    # Sort by date
    team_data.sort(key=lambda x: x['date'])
    periods = []
    current_period = None
    for entry in team_data:
        if current_period is None:
            current_period = {
                'team_id': entry['team_id'],
                'name': entry['name'],
                'start_date': entry['date'],
                'end_date': entry['date']
            }
        elif entry['name'] == current_period['name']:
            # Continue the period
            current_period['end_date'] = entry['date']
        else:
            # Team composition changed, close current period and start new
            periods.append(current_period)
            current_period = {
                'team_id': entry['team_id'],
                'name': entry['name'],
                'start_date': entry['date'],
                'end_date': entry['date']
            }
    if current_period:
        periods.append(current_period)
    return periods

def split_team_into_members(team_name):
    """Split a team name like 'Orla Shelly & Julie Roccati' into individual members"""
    if '&' in team_name:
        # Split by '&' and clean up each name
        members = [member.strip() for member in team_name.split('&')]
        return members
    else:
        # Single person team
        return [team_name.strip()]

def main():
    # File path
    file_path = "MyHome Wages Macros.xlsm"
    
    # Sheets to exclude
    excluded_sheets = ["17 May 25", "23 May 25", "30 May 25", "25 April 25", "6 June 25", "9 May 25"]
    
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    print("All sheet names:")
    for sheet_name in workbook.sheetnames:
        print(f"  '{sheet_name}'")
    print()
    
    all_team_data = []
    
    # Process each sheet
    for sheet_name in workbook.sheetnames:
        if sheet_name in excluded_sheets:
            continue
        
        print(f"Processing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        
        # Debug: print first 5 raw rows for a specific sheet
        if sheet_name == "06 Dec 24":
            print("First 5 raw rows from '06 Dec 24':")
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                print(row)
                if i >= 4:
                    break
        
        # Get data from the sheet
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Check if row has enough columns
            if len(row) < 13:
                continue
            if not row[0] or not row[1] or not row[12]:  # Skip if date, team, or team_id is empty
                continue
            date_str = str(row[0])
            team = str(row[1])
            team_id = str(row[12])  # Column M (index 12)
            
            # Skip if team_id is empty, contains formula, or is not a valid integer
            if not team_id or team_id.startswith('='):
                if sheet_name == "06 Dec 24" and len(all_team_data) < 5:
                    print(f"  Skipping due to team_id check: team_id='{team_id}'")
                continue
            try:
                int_team_id = int(team_id)
            except ValueError:
                if sheet_name == "06 Dec 24" and len(all_team_data) < 5:
                    print(f"  Skipping due to non-integer team_id: team_id='{team_id}'")
                continue
            team_id = str(int_team_id)  # Normalize to string integer
            
            # Debug: print raw values for first few rows
            if sheet_name == "06 Dec 24" and len(all_team_data) < 5:
                print(f"  Raw values: date={row[0]} ({type(row[0])}), team={row[1]} ({type(row[1])}), team_id={row[12]} ({type(row[12])})")
                print(f"  Processed: date_str='{date_str}', team='{team}', team_id='{team_id}'")
            
            # Parse date
            parsed_date = parse_date(date_str)
            if not parsed_date:
                if sheet_name == "06 Dec 24" and len(all_team_data) < 5:
                    print(f"  Skipping due to date parse: date_str='{date_str}'")
                continue
            
            # Clean team name
            team = team.strip()
            if not team:
                if sheet_name == "06 Dec 24" and len(all_team_data) < 5:
                    print(f"  Skipping due to empty team name")
                continue
            
            # Debug: print first few entries
            if len(all_team_data) < 5:
                print(f"  Found entry: date={date_str}, team={team}, team_id={team_id}")
            
            all_team_data.append({
                'date': parsed_date,
                'team_id': team_id,
                'name': team
            })
    
    # Group by team_id and create periods
    team_periods = []
    
    # Group data by team_id
    team_groups = {}
    for entry in all_team_data:
        team_id = entry['team_id']
        if team_id not in team_groups:
            team_groups[team_id] = []
        team_groups[team_id].append(entry)
    
    # Process each team_id group
    for team_id, entries in team_groups.items():
        periods = group_team_periods_allowing_gaps(entries)
        team_periods.extend(periods)
    
    # Sort by team_id, then by start_date
    team_periods.sort(key=lambda x: (int(x['team_id']), x['start_date']))
    
    # Split periods into individual staff member rows
    individual_periods = []
    for period in team_periods:
        team_members = split_team_into_members(period['name'])
        for member in team_members:
            individual_periods.append({
                'team_id': period['team_id'],
                'name': member,
                'original_team': period['name'],
                'start_date': period['start_date'],
                'end_date': period['end_date']
            })
    
    # Write to CSV
    output_file = "team_id_tracker.csv"
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = ['team_id', 'name', 'original_team', 'start_date', 'end_date']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        
        writer.writeheader()
        for period in individual_periods:
            writer.writerow({
                'team_id': period['team_id'],
                'name': period['name'],
                'original_team': period['original_team'],
                'start_date': period['start_date'].strftime('%d/%m/%Y'),
                'end_date': period['end_date'].strftime('%d/%m/%Y')
            })
    
    print(f"Generated {len(individual_periods)} individual staff periods in {output_file}")
    print(f"Total entries processed: {len(all_team_data)}")

if __name__ == "__main__":
    main() 