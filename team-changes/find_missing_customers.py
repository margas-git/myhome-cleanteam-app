import pandas as pd
import openpyxl
from datetime import datetime, time, timedelta

def parse_date(date_val):
    if isinstance(date_val, datetime):
        return date_val
    elif isinstance(date_val, str):
        try:
            return datetime.strptime(date_val, '%d/%m/%Y')
        except ValueError:
            try:
                return datetime.strptime(date_val, '%Y-%m-%d')
            except ValueError:
                return None
    return None

def main():
    # Load customer details for mapping
    customer_details = pd.read_excel("source_customer_details.xlsx")
    reference_customers = {str(row['name']).strip().lower(): str(row['name']).strip() for _, row in customer_details.iterrows()}
    
    print(f"Reference customers loaded: {len(reference_customers)}")
    
    # Load the workbook
    file_path = "MyHome Wages Macros app.xlsm"
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    # Sheets to exclude
    excluded_sheets = ["Totals", "Parameters", "Active Jobs", "17 May 25", "23 May 25", "30 May 25", "25 April 25", "6 June 25", "9 May 25"]
    
    # Track all customers found in source
    source_customers = set()
    missing_customers = set()
    
    # Process each sheet
    for sheet_name in workbook.sheetnames:
        if sheet_name in excluded_sheets:
            continue
        
        print(f"\nProcessing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        
        # Get data from the sheet
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Check if row has enough columns
            if len(row) < 13:
                continue
                
            # Extract customer name
            customer_name = row[2]  # Column C - Client/Customer name
            
            # Skip if customer name is empty
            if not customer_name:
                continue
            
            # Clean customer name
            customer_name_clean = str(customer_name).strip()
            if not customer_name_clean:
                continue
            
            # Add to source customers
            source_customers.add(customer_name_clean)
            
            # Check if in reference
            if customer_name_clean.lower() not in reference_customers:
                missing_customers.add(customer_name_clean)
    
    print(f"\n{'='*50}")
    print(f"SUMMARY:")
    print(f"Total customers in source: {len(source_customers)}")
    print(f"Total customers in reference: {len(reference_customers)}")
    print(f"Missing customers: {len(missing_customers)}")
    print(f"{'='*50}")
    
    if missing_customers:
        print(f"\nMISSING CUSTOMERS (in source but not in reference):")
        print(f"{'='*50}")
        for customer in sorted(missing_customers):
            print(f"- {customer}")
    else:
        print("\n✅ All customers in source are found in reference file!")
    
    # Show some reference customers for comparison
    print(f"\n{'='*50}")
    print(f"SAMPLE REFERENCE CUSTOMERS (first 10):")
    print(f"{'='*50}")
    for i, customer in enumerate(sorted(reference_customers.values())[:10]):
        print(f"{i+1}. {customer}")

if __name__ == "__main__":
    main() 