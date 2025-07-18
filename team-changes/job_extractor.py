import pandas as pd
import openpyxl
import json
from datetime import datetime, time, timedelta
import csv
import os
import requests
import time as time_module
from dotenv import load_dotenv

# Load environment variables from parent directory
load_dotenv('../.env')

def parse_date(date_val):
    """Parse date value which could be datetime object or string in various formats"""
    if isinstance(date_val, datetime):
        return date_val
    elif isinstance(date_val, str):
        # Try different date formats
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(date_val, fmt)
            except:
                continue
    return None

def geocode_address(address):
    """Geocode an address using Google Maps API"""
    api_key = os.getenv('GOOGLE_MAPS_API_KEY')
    if not api_key:
        print("Warning: GOOGLE_MAPS_API_KEY not found in environment variables")
        return None, None
    
    try:
        response = requests.get(
            f"https://maps.googleapis.com/maps/api/geocode/json",
            params={
                'address': f"{address}, Australia",
                'key': api_key
            }
        )
        
        data = response.json()
        
        if data['status'] == 'OK' and data['results']:
            location = data['results'][0]['geometry']['location']
            return str(location['lat']), str(location['lng'])
        else:
            print(f"Geocoding failed for '{address}': {data['status']}")
            return None, None
    except Exception as e:
        print(f"Error geocoding '{address}': {e}")
        return None, None

def main():
    # File path
    file_path = "MyHome Wages Macros app.xlsm"
    
    # Load customer details for mapping
    customer_details = pd.read_excel("source_customer_details.xlsx")
    name_to_id = {str(row['name']).strip(): int(row['id']) for _, row in customer_details.iterrows()}
    
    # Load additional customer data sources
    customer_all_df = pd.read_excel("source_customer_details.xlsx", sheet_name='all')
    customer_regular_df = pd.read_excel("source_customer_details.xlsx", sheet_name='regular-customers-wins')
    customer_combined_df = pd.read_excel("source_customer_details.xlsx", sheet_name='combined')
    
    # Load users for staff mapping
    users_df = pd.read_csv("source_users.csv")
    staff_name_to_id = {}
    staff_name_to_full = {}
    for _, row in users_df.iterrows():
        first = str(row['first_name']).strip()
        last = str(row['last_name']).strip()
        full = f"{first} {last}".strip()
        staff_name_to_id[full.lower()] = str(row['id'])
        staff_name_to_full[full.lower()] = full
    
    # Sheets to exclude
    excluded_sheets = ["Totals", "Parameters", "Active Jobs", "17 May 25", "23 May 25", "30 May 25", "25 April 25", "6 June 25", "9 May 25"]
    
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    print("All sheet names:")
    for sheet_name in workbook.sheetnames:
        print(f"  '{sheet_name}'")
    print()
    
    all_job_data = []
    all_time_entries = []
    job_counter = 0  # Track job IDs as we create them
    
    # Process each sheet
    for sheet_name in workbook.sheetnames:
        if sheet_name in excluded_sheets:
            continue
        
        print(f"Processing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        
        # Group rows by customer to handle additional staff for jobs
        customer_jobs = {}
        
        # First pass: collect all rows for each customer
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Check if row has enough columns
            if len(row) < 13:
                continue
                
            # Extract required fields
            date_val = row[0]  # Column A - Date
            team_members = row[1]  # Column B - Team members
            customer_name = row[2]  # Column C - Client/Customer name
            start_time = row[3]  # Column D - Start time
            finish_time = row[4]  # Column E - Finish time
            lunch_break = row[5]  # Column F - Lunch Break (NEW)
            price = row[7]  # Column H - Quoted H
            team_id = row[12]  # Column M - Team ID
            
            # Skip if customer name is empty
            if not customer_name:
                continue
                
            # Skip if any other required field is empty
            if not date_val or not start_time:
                continue
            
            # Clean customer name
            customer_name_clean = str(customer_name).strip()
            if not customer_name_clean:
                continue
            
            # Create a unique key for this customer job (date + customer)
            job_key = f"{date_val}_{customer_name_clean}"
            
            if job_key not in customer_jobs:
                customer_jobs[job_key] = {
                    'date_val': date_val,
                    'customer_name': customer_name_clean,
                    'start_time': start_time,
                    'finish_time': finish_time,
                    'lunch_break': lunch_break,
                    'price': price,
                    'team_members': [],
                    'additional_staff': [],
                    'team_id': None
                }
            
            # Add team member to appropriate list
            if team_members:
                staff_member = str(team_members).strip()
                if staff_member:
                    if team_id and str(team_id).strip():  # Has team ID = core team member
                        customer_jobs[job_key]['team_members'].append(staff_member)
                        customer_jobs[job_key]['team_id'] = team_id
                    else:  # No team ID = additional staff
                        customer_jobs[job_key]['additional_staff'].append(staff_member)
        
        # Second pass: create jobs and time entries
        for job_key, job_data in customer_jobs.items():
            # Skip if no team members at all
            if not job_data['team_members'] and not job_data['additional_staff']:
                continue
                
            # Skip if team_id is not a valid integer
            try:
                team_id_str = str(job_data['team_id'])
                int_team_id = int(team_id_str)
            except (ValueError, TypeError):
                continue
                
            # Parse date
            parsed_date = parse_date(job_data['date_val'])
            if not parsed_date:
                continue
            
            # Map customer name to customer_id
            customer_id = name_to_id.get(job_data['customer_name'])
            if customer_id is None:
                print(f"Warning: Customer name '{job_data['customer_name']}' not found in source_customer_details.xlsx. Using customer_id = 0.")
                customer_id = 0  # Use 0 as default for missing customers
            
            # Combine date and start time
            if isinstance(job_data['start_time'], time):
                combined_start_datetime = datetime.combine(parsed_date.date(), job_data['start_time'])
            else:
                # If start_time is not a time object, use just the date
                combined_start_datetime = parsed_date
            
            # Combine date and finish time
            if isinstance(job_data['finish_time'], time):
                combined_finish_datetime = datetime.combine(parsed_date.date(), job_data['finish_time'])
            else:
                # If finish_time is not a time object, use just the date
                combined_finish_datetime = parsed_date
            
            # Convert to UTC by subtracting 10 hours
            utc_start_datetime = combined_start_datetime - timedelta(hours=10)
            utc_finish_datetime = combined_finish_datetime - timedelta(hours=10)
            
            # Process lunch break data
            lunch_break_value = ''
            if job_data['lunch_break'] is not None and str(job_data['lunch_break']).strip():
                lunch_break_value = str(job_data['lunch_break']).strip()
            
            # Calculate clock_out_lunch_break
            clock_out_lunch_break = ''
            if lunch_break_value.strip().lower() == 'yes':
                # If lunch_break is "Yes", calculate finish time minus 30 minutes
                if isinstance(job_data['finish_time'], time):
                    actual_finish_time = combined_finish_datetime
                    lunch_break_finish_time = actual_finish_time - timedelta(minutes=30)
                    utc_lunch_break_finish = lunch_break_finish_time - timedelta(hours=10)
                    clock_out_lunch_break = utc_lunch_break_finish.strftime('%Y-%m-%d %H:%M:%S+00')
                else:
                    lunch_break_finish_time = combined_finish_datetime - timedelta(minutes=30)
                    utc_lunch_break_finish = lunch_break_finish_time - timedelta(hours=10)
                    clock_out_lunch_break = utc_lunch_break_finish.strftime('%Y-%m-%d %H:%M:%S+00')
            elif lunch_break_value.strip().lower() == 'no':
                # If lunch_break is "No", use the finish time as is
                if isinstance(job_data['finish_time'], time):
                    actual_finish_time = combined_finish_datetime
                    utc_finish_time = actual_finish_time - timedelta(hours=10)
                    clock_out_lunch_break = utc_finish_time.strftime('%Y-%m-%d %H:%M:%S+00')
                else:
                    utc_finish_time = combined_finish_datetime - timedelta(hours=10)
                    clock_out_lunch_break = utc_finish_time.strftime('%Y-%m-%d %H:%M:%S+00')
            
            # If price is empty, set to 0
            if job_data['price'] is None or job_data['price'] == "":
                price = 0
            else:
                price = job_data['price']
            
            # Increment job counter
            job_counter += 1
            
            all_job_data.append({
                'id': job_counter,
                'customer_id': customer_id,
                'team_id': str(int_team_id),
                'status': 'completed',
                'created_at': utc_start_datetime,
                'price': price,
                'customer_name': job_data['customer_name'],
                'team_members_at_creation': json.dumps(job_data['team_members']),
                'additional_staff': json.dumps(job_data['additional_staff'])
            })
            
            # Add time entry for each staff member who worked on this job (both team and additional)
            all_staff = job_data['team_members'] + job_data['additional_staff']
            for staff_group in all_staff:
                # Split staff names by '&' if multiple staff in one entry
                if '&' in staff_group:
                    individual_staff = [member.strip() for member in staff_group.split('&') if member.strip()]
                else:
                    individual_staff = [staff_group.strip()]
                
                # Create a time entry for each individual staff member
                for staff_member in individual_staff:
                    staff_key = staff_member.lower()
                    user_id = staff_name_to_id.get(staff_key, '')
                    all_time_entries.append({
                        'id': len(all_time_entries) + 1,
                        'user_id': user_id,
                        'staff': staff_member,
                        'job_id': job_counter,
                        'clock_in_time': utc_start_datetime,
                        'clock_out_time': clock_out_lunch_break,  # Use clock_out_lunch_break value
                        'lunch_break': '',  # Remove lunch_break values
                        'geofence_override': '',
                        'auto_lunch_deducted': ''
                    })
    
    # Sort by created_at, then by customer_id, then by team_id
    all_job_data.sort(key=lambda x: (x['created_at'], x['customer_id'], int(x['team_id'])))
    
    # Reassign IDs starting at 1 for the sorted jobs
    job_id_mapping = {}
    for i, job in enumerate(all_job_data):
        old_id = job['id']
        new_id = i + 1
        job['id'] = new_id
        job_id_mapping[old_id] = new_id
    
    # Update time entries with new job_ids
    for time_entry in all_time_entries:
        if time_entry['job_id'] in job_id_mapping:
            time_entry['job_id'] = job_id_mapping[time_entry['job_id']]
    
    # Collect unique customers from job data (MyHome Wages spreadsheet)
    # This ensures we only include customers who have completed work
    unique_customers = set()
    
    # Add customers from job data (MyHome Wages spreadsheet)
    for job in all_job_data:
        unique_customers.add(job['customer_name'])
    
    # Add additional customers from combined sheet based on criteria:
    # - Column M (active) = TRUE
    # - Column N (created_at) >= 13/06/2025
    target_date = pd.to_datetime('13/06/2025').tz_localize('UTC')
    
    # Parse dates correctly (format is YY-MM-DD)
    def parse_date_correctly(date_str):
        if pd.isna(date_str):
            return pd.NaT
        try:
            # Handle format like "25-06-26 04:00:00+00" (YY-MM-DD)
            if isinstance(date_str, str) and len(date_str) >= 8:
                year = int('20' + date_str[:2])  # Convert YY to 20YY
                month = int(date_str[3:5])
                day = int(date_str[6:8])
                return pd.Timestamp(year, month, day).tz_localize('UTC')
            else:
                return pd.to_datetime(date_str)
        except:
            return pd.NaT
    
    # Apply correct date parsing
    customer_combined_df['parsed_created_at'] = customer_combined_df['created_at'].apply(parse_date_correctly)
    
    additional_customers = customer_combined_df[
        (customer_combined_df['active'] == True) & 
        (customer_combined_df['parsed_created_at'] >= target_date)
    ]
    
    for _, row in additional_customers.iterrows():
        unique_customers.add(str(row['name']).strip())
    
    # Create customers data
    all_customers_data = []
    customer_counter = 1
    
    for customer_name in sorted(unique_customers):
        # Find customer in source data
        customer_id = name_to_id.get(customer_name, 0)
        
        # Get data from 'all' sheet (case-insensitive matching)
        all_row = customer_all_df[customer_all_df['Customer'].str.lower() == customer_name.lower()]
        if len(all_row) == 0:
            # Try partial matching
            all_row = customer_all_df[customer_all_df['Customer'].str.lower().str.contains(customer_name.lower(), na=False)]
        
        # Get data from 'combined' sheet (case-insensitive matching)
        combined_row = customer_combined_df[customer_combined_df['name'].str.lower() == customer_name.lower()]
        if len(combined_row) == 0:
            # Try partial matching
            combined_row = customer_combined_df[customer_combined_df['name'].str.lower().str.contains(customer_name.lower(), na=False)]
        
        # Get active status from combined sheet Column M
        active_status = combined_row['active'].iloc[0] if len(combined_row) > 0 else False
        
        # Get price - if customer is from combined sheet, use combined sheet Column H, otherwise use job data
        if len(combined_row) > 0:
            price = combined_row['price'].iloc[0] if len(combined_row) > 0 else 0
        else:
            # Get price from job data (Wages app.xlsx Column H)
            price = 0
            for job in all_job_data:
                if job['customer_name'].lower() == customer_name.lower():
                    price = job['price']
                    break
        
        # Get clean_frequency from regular-customers-wins sheet Column G
        regular_row = customer_regular_df[customer_regular_df['Customer Name'].str.lower() == customer_name.lower()]
        if len(regular_row) == 0:
            # Try partial matching
            regular_row = customer_regular_df[customer_regular_df['Customer Name'].str.lower().str.contains(customer_name.lower(), na=False)]
        clean_frequency = str(regular_row['Frequency'].iloc[0]) if len(regular_row) > 0 else "One off"
        
        # Clean up frequency values
        if clean_frequency.lower() == '3weekly':
            clean_frequency = 'Tri-weekly'
        elif clean_frequency.lower() == '6weekly':
            clean_frequency = 'One off'
        
        # Use combined sheet data if available, otherwise fall back to all sheet
        if len(combined_row) > 0:
            address = str(combined_row['address'].iloc[0]) if len(combined_row) > 0 else ''
            latitude = ''  # Set to blank as requested
            phone_raw = str(combined_row['phone'].iloc[0]) if len(combined_row) > 0 else ''
            # Get created_at from all sheet Column F
            created_at = str(all_row.iloc[0, 5]) if len(all_row) > 0 else ''  # Column F (index 5)
        else:
            # Fall back to all sheet data
            address = str(all_row['Primary Address'].iloc[0]) if len(all_row) > 0 else ''
            latitude = ''  # Set to blank as requested
            phone_raw = str(all_row['Phone No.'].iloc[0]) if len(all_row) > 0 else ''
            created_at = str(all_row.iloc[0, 5]) if len(all_row) > 0 else ''  # Column F (index 5)
        
        # Add leading "0" to phone numbers and remove trailing .0
        phone_raw_str = str(phone_raw) if phone_raw else ''
        phone = '0' + phone_raw_str if phone_raw_str and not phone_raw_str.startswith('0') else phone_raw_str
        phone = phone.replace('.0', '')  # Remove trailing .0
        
        # Set other fields as requested
        email = ''  # Set to blank as requested
        notes = ''  # Set to blank as requested
        target_time_minutes = ''  # Set to blank as requested
        average_wage_ratio = ''  # Set to blank as requested
        
        # Set latitude and longitude to empty (disabled geocoding)
        latitude = ''
        longitude = ''
        
        # Format created_at date from Column F to YYYY-MM-DD 04:00:00+00 format
        try:
            # Parse date in format like "6th Sep 24"
            if isinstance(created_at, str) and created_at.strip():
                # Use pandas to parse the date with dayfirst=True
                date_obj = pd.to_datetime(created_at, dayfirst=True)
                year = date_obj.strftime('%Y')  # 4-digit year
                month = date_obj.strftime('%m')  # 2-digit month
                day = date_obj.strftime('%d')    # 2-digit day
                formatted_date = f'{year}-{month}-{day} 04:00:00+00'
                created_at = formatted_date
            else:
                created_at = '2025-01-01 04:00:00+00'
        except:
            created_at = '2025-01-01 04:00:00+00'  # Default if parsing fails
        
        all_customers_data.append({
            'id': customer_counter,
            'name': customer_name,
            'address': address,
            'latitude': latitude,  # Now geocoded from address
            'longitude': longitude,  # Now geocoded from address
            'phone': phone,  # Now with leading "0"
            'email': email,  # Now blank as requested
            'price': price,  # Now from job data
            'clean_frequency': clean_frequency,  # Now from regular-customers-wins sheet
            'notes': notes,  # Now blank as requested
            'target_time_minutes': target_time_minutes,  # Now blank as requested
            'average_wage_ratio': average_wage_ratio,  # Now blank as requested
            'is_friends_family': False,  # Static FALSE
            'friends_family_minutes': '',  # Blank as specified
            'active': active_status,  # Now from combined sheet
            'created_at': created_at
        })
        customer_counter += 1
    
    # Sort customers by created_at date (ascending) and reassign IDs
    all_customers_data.sort(key=lambda x: x['created_at'])
    
    # Create mapping from old customer names to new IDs
    customer_name_to_new_id = {}
    for i, customer in enumerate(all_customers_data):
        old_id = customer['id']
        new_id = i + 1
        customer['id'] = new_id
        customer_name_to_new_id[customer['name']] = new_id
    
    # Update job customer IDs to match the new customer IDs
    for job in all_job_data:
        customer_name = job['customer_name']
        if customer_name in customer_name_to_new_id:
            job['customer_id'] = customer_name_to_new_id[customer_name]
        else:
            job['customer_id'] = 0  # Default for missing customers
    
    # Create a new Excel workbook
    output_file = "MyHome_Data.xlsx"
    
    # Create DataFrame for jobs
    jobs_df = pd.DataFrame(all_job_data)
    
    # Convert created_at column to string format for Excel
    jobs_df['created_at'] = pd.to_datetime(jobs_df['created_at']).dt.strftime('%Y-%m-%d %H:%M:%S+00')
    
    # Ensure price column is never empty (should already be handled above, but just in case)
    jobs_df['price'] = jobs_df['price'].apply(lambda x: 0 if x is None or x == "" else x)
    
    # Create DataFrame for time entries
    time_entries_df = pd.DataFrame(all_time_entries)
    
    # Convert time columns to string format for Excel
    time_entries_df['clock_in_time'] = pd.to_datetime(time_entries_df['clock_in_time']).dt.strftime('%Y-%m-%d %H:%M:%S+00')
    time_entries_df['clock_out_time'] = pd.to_datetime(time_entries_df['clock_out_time']).dt.strftime('%Y-%m-%d %H:%M:%S+00')
    
    # Sort time entries by clock_in_time from oldest to newest
    time_entries_df = time_entries_df.sort_values('clock_in_time', ascending=True)
    
    # Reassign time entry IDs starting at 1 for the sorted entries
    time_entries_df['id'] = range(1, len(time_entries_df) + 1)
    
    # Create DataFrame for customers
    customers_df = pd.DataFrame(all_customers_data)
    
    # Write to Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        jobs_df.to_excel(writer, sheet_name='jobs', index=False)
        time_entries_df.to_excel(writer, sheet_name='time_entries', index=False)
        customers_df.to_excel(writer, sheet_name='customers', index=False)
    
    print(f"Generated {len(all_job_data)} job entries in {output_file}")
    print(f"Generated {len(all_time_entries)} time entries in {output_file}")
    print(f"Generated {len(all_customers_data)} customer entries in {output_file}")
    print(f"Total entries processed: {len(all_job_data)}")
    print(f"Excel file '{output_file}' created with 'jobs', 'time_entries', and 'customers' sheets")
    print("You can now add more sheets to this Excel file as needed.")

if __name__ == "__main__":
    main() 